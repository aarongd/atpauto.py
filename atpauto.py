import openpyxl

#Where information pulled from the Action Tracker, SLA Tracker, and cost org will go.
li = {
    "Candidate Name" : "",
    "Company" : "",
    "JITR" : "",
    "CSR" : "",
    "Labor Category" : "",
    "Level" : "",
    "Effective Date" : "",
    "Submitted Rate to CACI" : "",
    "SLA" : "",
    "CLIN" : "0820",
    "Resource ID" : "",
    "Cost Center" : "",
}

#the two inputs needed to put in manually
csr_input = input("CSR Number: ")
#resource_id_input = input("Resource ID: ")

#li["Resource ID"] = resource_id_input

#opens the Action Tracker workbook
path = "C:\\Users\\Aaron\\Desktop\\action tracker.xlsx"
wb = openpyxl.load_workbook(path)
sheets = wb.sheetnames
ws = wb.active
#n=0
#ws = wb[sheets[n]].active

#to find all rows within the CSR column in the Action tab
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        if cell.value == csr_input:
            #saving all relevant values from this excel book to memory
            full_name = ws.cell(row=cell.row, column=6).value
            company = ws.cell(row=cell.row, column=9).value
            jitr = ws.cell(row=cell.row, column=2).value
            labor_category = ws.cell(row=cell.row, column=7).value
            level = ws.cell(row=cell.row, column=8).value
            start_date = ws.cell(row=cell.row, column=18).value
            rate2caci = ws.cell(row=cell.row, column=10).value
            #start adding all saved values from this excel book to the list
            li["Candidate Name"] = full_name
            li["Company"] = company
            li["JITR"] = jitr
            li["CSR"] = csr_input
            li["Labor Category"] = labor_category
            li["Level"] = level
            li["Effective Date"] = start_date
            li["Submitted Rate to CACI"] = rate2caci
            # Cost center is a constant unless for a few specific JITRs
            if jitr in (1124, 1125, 1126, 1158, 1160, 1166):
                li["Cost Center"] = 3373
            else:
                li["Cost Center"] = 3393
        else:
            print("CSR value not found")

#opens the SLA tracker
path2 = "C:\\Users\\Aaron\\Desktop\\sla tracker.xlsx"
wb2 = openpyxl.load_workbook(path2)
sheets2 = wb2.sheetnames
ws2 =wb2.active
#n=0
#ws = wb[sheets[n]].active

for row in ws2.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        jitr_value = li['JITR']
        if cell.value == int(jitr_value):
            #saving all relevant values from this excel book to memory
            sla = ws2.cell(row=cell.row, column=2).value
            #start adding all saved values from this excel book to the list
            li["SLA"] = sla


print(li)

